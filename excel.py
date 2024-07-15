import openpyxl
from openpyxl.styles import PatternFill

def write_to_excel(data, period: list[str]):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Monthly Report for {period}"

    header = ["Наименование товара", "Цвет товара", "Количество", "Ед. измерения", "Дата и время выдачи товара", "Примечание"]
    header_fill = PatternFill(patternType="solid", start_color="00FF00", end_color="00FF00")
    sheet.append(header)
    for cell in sheet[1]:
        cell.fill = header_fill

    for i in data:
        for j in i:
            sheet.append(j)
            if j[5]:
                for cell in sheet[sheet.max_row]:
                    cell.fill = PatternFill(patternType="solid", start_color="FF0000", end_color="FF0000")
    
    return workbook